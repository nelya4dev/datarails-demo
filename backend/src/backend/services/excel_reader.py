"""Excel file reader for processing uploaded .xlsx files."""

from pathlib import Path
from typing import List, Dict, Optional, Any
from datetime import datetime, date
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """Read and parse Excel files with validation and type conversion.

    Handles reading Excel files (.xlsx, .xlsm) and converting them to
    structured data with proper type handling for dates, numbers, etc.

    Usage:
        reader = ExcelReader("path/to/file.xlsx")
        sheets = reader.get_sheet_names()
        data = reader.read_sheet("Employees")
    """

    def __init__(self, file_path: str):
        """Initialize Excel reader.

        Args:
            file_path: Path to Excel file

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file is not a valid Excel file
        """
        self.file_path = Path(file_path)

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if not self.file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Invalid Excel file format: {self.file_path.suffix}")

        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")

    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook.

        Returns:
            List of sheet names

        Example:
            >>> reader.get_sheet_names()
            ['Employees', 'Projects']
        """
        return self.workbook.sheetnames

    def has_sheet(self, sheet_name: str) -> bool:
        """Check if sheet exists in workbook.

        Args:
            sheet_name: Name of sheet to check

        Returns:
            True if sheet exists, False otherwise
        """
        return sheet_name in self.workbook.sheetnames

    def read_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Read sheet and return list of dictionaries.

        First row is treated as headers. Subsequent rows are data.
        Automatically converts Excel dates to Python date objects.

        Args:
            sheet_name: Name of sheet to read

        Returns:
            List of dictionaries, one per row

        Raises:
            ValueError: If sheet doesn't exist or has no data

        Example:
            >>> data = reader.read_sheet("Employees")
            >>> data[0]
            {
                'employee_id': 'E0001',
                'name': 'Kevin Davis',
                'department_code': 'DEV',
                'salary': 78289.0,
                'hire_date': date(2023, 4, 5)
            }
        """
        if not self.has_sheet(sheet_name):
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {', '.join(self.get_sheet_names())}"
            )

        sheet = self.workbook[sheet_name]

        # Get headers from first row
        headers = self._get_headers(sheet)

        if not headers:
            raise ValueError(f"Sheet '{sheet_name}' has no headers in first row")

        # Read data rows
        data = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            # Create dictionary from headers and values
            row_dict = {}
            for header, value in zip(headers, row):
                if header:  # Skip None headers
                    row_dict[header] = self._clean_cell_value(value)

            # Add row number for error tracking
            row_dict['_excel_row_number'] = row_num

            data.append(row_dict)

        return data

    def _get_headers(self, sheet: Worksheet) -> List[Optional[str]]:
        """Extract headers from first row.

        Args:
            sheet: Worksheet object

        Returns:
            List of header names (strings or None)
        """
        headers = []
        for cell in sheet[1]:
            value = cell.value
            if value is not None:
                # Convert to string and clean
                header = str(value).strip()
                headers.append(header if header else None)
            else:
                headers.append(None)
        return headers

    def _clean_cell_value(self, value: Any) -> Any:
        """Clean and convert cell value to appropriate Python type.

        Handles:
        - Excel dates → Python date objects
        - Excel datetime → Python date objects (date only)
        - Strings → Stripped strings
        - Numbers → Keep as-is
        - None → None

        Args:
            value: Raw cell value from openpyxl

        Returns:
            Cleaned value with appropriate type
        """
        # Handle None
        if value is None:
            return None

        # Handle datetime objects (Excel stores dates as datetime)
        if isinstance(value, datetime):
            return value.date()  # Convert to date only

        # Already a date object
        if isinstance(value, date):
            return value

        # Handle strings
        if isinstance(value, str):
            stripped = value.strip()
            return stripped if stripped else None

        # Numbers, booleans, etc. - keep as-is
        return value

    def get_row_count(self, sheet_name: str) -> int:
        """Get number of data rows in sheet (excluding header).

        Args:
            sheet_name: Name of sheet

        Returns:
            Number of data rows
        """
        if not self.has_sheet(sheet_name):
            raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet = self.workbook[sheet_name]
        # max_row includes header, so subtract 1
        return max(0, sheet.max_row - 1)

    def get_column_names(self, sheet_name: str) -> List[str]:
        """Get column names (headers) from sheet.

        Args:
            sheet_name: Name of sheet

        Returns:
            List of column names
        """
        if not self.has_sheet(sheet_name):
            raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet = self.workbook[sheet_name]
        headers = self._get_headers(sheet)
        return [h for h in headers if h is not None]

    def validate_required_sheets(self, required_sheets: List[str]) -> None:
        """Validate that all required sheets exist.

        Args:
            required_sheets: List of required sheet names

        Raises:
            ValueError: If any required sheet is missing
        """
        missing_sheets = [
            sheet for sheet in required_sheets
            if not self.has_sheet(sheet)
        ]

        if missing_sheets:
            raise ValueError(
                f"Missing required sheets: {', '.join(missing_sheets)}. "
                f"Available sheets: {', '.join(self.get_sheet_names())}"
            )

    def validate_required_columns(
            self,
            sheet_name: str,
            required_columns: List[str]
    ) -> None:
        """Validate that sheet has all required columns.

        Args:
            sheet_name: Name of sheet to validate
            required_columns: List of required column names

        Raises:
            ValueError: If any required column is missing
        """
        actual_columns = set(self.get_column_names(sheet_name))
        required_set = set(required_columns)
        missing_columns = required_set - actual_columns

        if missing_columns:
            raise ValueError(
                f"Sheet '{sheet_name}' missing required columns: "
                f"{', '.join(sorted(missing_columns))}. "
                f"Available columns: {', '.join(sorted(actual_columns))}"
            )

    def close(self) -> None:
        """Close the workbook and release resources."""
        if hasattr(self, 'workbook'):
            self.workbook.close()

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - closes workbook."""
        self.close()

    def __repr__(self) -> str:
        return (
            f"<ExcelReader "
            f"file='{self.file_path.name}' "
            f"sheets={len(self.get_sheet_names())}>"
        )
